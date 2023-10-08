import pandas as pd
import numpy as np
import os
import itertools
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def listFiles(path):
    '''
        根据指定的路径,遍历所有后缀名是xlsx的文件,这个方法只遍历了当前目录,并没有迭代遍历当前路径下的子目录
    '''
    files_list = []
    for file in os.listdir(path):
        print(file)
        #print(type(file))
        #~开头的文件是临时文件
        if(file.startswith('~')):
            continue
        if os.path.splitext(file)[1] == '.xlsx':
            files_list.append(file)
    print("---------valid file start---------------")
    print(files_list)
    print("---------valid file end-----------------")
    return files_list

def main():
    #for filename in ["933.xlsx","369.xlsx"]:
    for filename in ["369.xlsx"]:
    #for filename in listFiles('.'):
        df = pd.read_excel(filename, dtype=str, engine='openpyxl')
        df.head()

        #margin这个参数表示时候显示总计的这一栏
        pt = pd.pivot_table(df, index=["MAWB","Container No.","HAWB"], values=["CBP Status"], aggfunc=["count"], margins=True)
        #pt = pd.pivot_table(df, index=["MAWB","Container No.","HAWB"], values=["CBP Status"], aggfunc=["count"], margins=False)
        

        #print(pt)
        #print(type(pt))

        #pt.to_excel("../369-pivot.xlsx")
        
        #https://cloud.tencent.com/developer/article/1770494
        #writer = pd.ExcelWriter("../"+filename, engine='openpyxl',mode='a', if_sheet_exists='replace')
        writer = pd.ExcelWriter("../zx-pivot-"+filename, engine='openpyxl',mode='w')
        pt.to_excel(writer, sheet_name="zx-pivot-table")
        workbook = writer.book
        worksheet = writer.sheets['zx-pivot-table']
        #print(workbook)

        #设定边框的样式
        side = Side(style="thin")
        border = Border(left=side, right=side, top=side, bottom=side)

        #为所有单元格设置边框
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border


        final_col_width = []
        for column in worksheet.iter_cols():
            col_width = []
            for cell in column:
                #print(type(cell))
                #print(cell.value)
                #print("column width=",len(str(cell.value)) if cell.value is not None else 0)
                col_width.append(len(str(cell.value)) if cell.value is not None else 0)
            #print("****************************")
            final_col_width.append(max(col_width))

        #print("final_col_width=", final_col_width)
        #print(help(worksheet))
        #设定column的宽度,这里可以想办法把大写字母ABCD转换为对应的ASCII码值,就变成整数了,然后可以用循环来遍历所有的列,
        worksheet.column_dimensions["A"].width = final_col_width[0] + 5
        worksheet.column_dimensions["B"].width = final_col_width[1] + 5
        worksheet.column_dimensions["C"].width = final_col_width[2] + 5
        worksheet.column_dimensions["D"].width = final_col_width[3]

        '''
        print(worksheet.max_column)
        for i in range(len(final_col_width)):
            worksheet.column_dimensions
        '''
        writer.close()

        '''
        设置header
        &[Date]
        &[Page]
        &[Pages]
        >>> from openpyxl.workbook import Workbook
        >>>
        >>> wb = Workbook()
        >>> ws = wb.active
        >>>
        >>> ws.oddHeader.left.text = "Page &[Page] of &N"
        >>> ws.oddHeader.left.size = 14
        >>> ws.oddHeader.left.font = "Tahoma,Bold"
        >>> ws.oddHeader.left.color = "CC3366"

        https://zhuanlan.zhihu.com/p/459295902

        #设置打印区域
        >>> from openpyxl.workbook import Workbook
        >>>
        >>> wb = Workbook()
        >>> ws = wb.active
        >>>
        >>> ws.print_area = 'A1:F10'
        '''

        """
        with pd.ExcelWriter("933.xlsx") as writer:
            pt.to_excel("../933-pivot.xlsx", sheet_name="zx-pivot-table")
            worksheet = writer.sheets['zx-pivot-table']
            for i in range(1,3):
                worksheet.set_column(i, i, 20)

            writer.save()
        """

    '''
    本文代码及数据集来自《超简单：用Python让Excel飞起来（实战150例）》
    # 打印一个工作簿中的所有工作表
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('各月销售数量表.xlsx')
    workbook.api.PrintOut(Copies=2, ActivePrinter='DESKTOP-HP01', Collate=True) # 打印工作簿中的所有工作表，这里指定打印份数为两份，打印机为“DESKTOP-HP01”
    workbook.close()
    app.quit()

    # 打印一个工作簿中的一个工作表
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('各月销售数量表.xlsx')
    worksheet = workbook.sheets['1月']
    worksheet.api.PrintOut(Copies=2, ActivePrinter='DESKTOP-HP01', Collate=True)
    workbook.close()
    app.quit()

    # 打印多个工作簿
    from pathlib import Path
    import xlwings as xw
    folder_path = Path('各地区销售数量')
    file_list = folder_path.glob('*.xls*')
    app = xw.App(visible=False, add_book=False)
    for i in file_list:
        workbook = app.books.open(i)
        workbook.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
        workbook.close()
    app.quit()

    # 打印多个工作簿中的同名工作表
    from pathlib import Path
    import xlwings as xw
    folder_path = Path('各地区销售数量')
    file_list = folder_path.glob('*.xls*')
    app = xw.App(visible=False, add_book=False)
    for i in file_list:
        workbook = app.books.open(i)
        worksheet = workbook.sheets['销售数量']
        worksheet.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
        workbook.close()
    app.quit()

    # 打印工作表的指定单元格区域
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('销售表.xlsx')
    worksheet = workbook.sheets['总表']
    area = worksheet.range('A1:I10')
    area.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
    workbook.close()
    app.quit()

    # 按指定的缩放比例打印工作表
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('销售表.xlsx')
    worksheet = workbook.sheets['总表']
    worksheet.api.PageSetup.Zoom = 80 # 按工作表原始大小的80%进行打印
    worksheet.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
    workbook.close()
    app.quit()

    # 在纸张的居中位置打印工作表
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('各月销售数量表.xlsx')
    worksheet = workbook.sheets['1月']
    worksheet.api.PageSetup.CenterHorizontally = True # 调整工作表在纸张上的水平位置
    worksheet.api.PageSetup.CenterVertically = True # 调整工作表在纸张上的垂直位置
    worksheet.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
    workbook.close()
    app.quit()

    # 打印工作表时打印行号和列号
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('各月销售数量表.xlsx')
    worksheet = workbook.sheets['1月']
    worksheet.api.PageSetup.PrintHeadings = True
    worksheet.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
    workbook.close()
    app.quit()

    # 重复打印工作表的标题行
    import xlwings as xw
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open('销售表.xlsx')
    worksheet = workbook.sheets['总表']
    worksheet.api.PageSetup.PrintTitleRows = '$1:$1' # 将工作表第一行设置为要重复打印的标题行
    worksheet.api.PageSetup.Zoom = 55 # 按工作表原始大小的55%进行打印
    worksheet.api.PrintOut(Copies=1, ActivePrinter='DESKTOP-HP01', Collate=True)
    workbook.close()
    app.quit()

    '''

main()